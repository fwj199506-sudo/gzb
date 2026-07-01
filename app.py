import streamlit as st
import pandas as pd
import numpy as np
import io
import re
import os

# ================= 页面配置 =================
st.set_page_config(page_title="私募资管产品估值表分析系统", layout="wide")
st.title("📊 私募资管产品估值表精算系统 (底层数组坐标版)")
st.markdown("采用二维数组坐标级定位技术，严格遵循 CAS 科目代码层级分类，解决列名错位与分类混淆问题。")

# ================= 工具函数 =================
def parse_num(val):
    """安全地将各种格式转换为浮点数"""
    if pd.isna(val):
        return 0.0
    try:
        val_str = str(val).replace(',', '').replace(' ', '').replace('，', '').strip()
        if not val_str or val_str in ['-', '--', 'None', 'nan']:
            return 0.0
        if val_str.endswith('%'):
            return float(val_str[:-1]) / 100.0
        return float(val_str)
    except:
        return 0.0

def norm_text(x):
    """统一文本格式：去空格、去全角空格、去首尾横线"""
    s = str(x).replace('\u3000', '').replace(' ', '').strip()
    s = re.sub(r'^[－\-—]+', '', s)
    return s

def clean_code_str(raw_code):
    """清洗证券代码，如 149407 SZ -> 149407.SZ；1103.06.01.196519 SH -> 1103.06.01.196519.SH"""
    code_str = str(raw_code).strip()
    if not code_str or code_str in ['nan', 'None']:
        return ""
    code_str = re.sub(r'\s+([A-Za-z]{2,4})$', r'.\1', code_str)
    return code_str

def extract_valuation_date(file_name):
    """从估值表文件名提取 YYYYMMDD 估值日期。"""
    date_match = re.search(r'_(\d{8})_', os.path.basename(str(file_name)))
    return date_match.group(1) if date_match else ""

def build_output_filename(uploaded_files):
    """根据上传文件估值日期生成下载文件名，支持日期范围。"""
    dates = sorted({
        extract_valuation_date(getattr(file, "name", ""))
        for file in uploaded_files
        if extract_valuation_date(getattr(file, "name", ""))
    })
    if not dates:
        suffix = ""
    elif len(dates) == 1:
        suffix = f"_{dates[0]}"
    else:
        suffix = f"_{dates[0]}-{dates[-1]}"
    return f"私募产品多维度估值透视表{suffix}.xlsx"

def is_accrued_interest_row(code_str):
    """
    判断是否为应计利息子行（避免与本金行重复计算）。
    债券估值表中每只券拆为本金(.01)和应计利息(.03)两行，
    应计利息行的第三层 CAS 科目为 03，如 1103.06.03.196519.SH。
    """
    c = clean_code_str(code_str)
    parts = c.split('.')
    return len(parts) >= 4 and parts[2] == '03'

def extract_ticker(full_code):
    """
    从完整 CAS 科目代码中提取纯证券代码。
    如 1103.02.01.122498.SH → 122498.SH
       1102.34.01.301308.SZ → 301308.SZ
       1108.01.01.WX6ZXT.OTC → WX6ZXT.OTC
    CAS 层级段均为短代码（≤4位纯数字，或 B1 这类单字母+数字），
    证券代码是第一个打破该规则的段及其后续部分。
    """
    code_str = clean_code_str(full_code)
    if not code_str:
        return ""
    parts = code_str.split('.')
    for i, p in enumerate(parts):
        # CAS 层级：≤4 位纯数字，或单字母 + 可选 1 位数字（如 B1）
        is_cas = len(p) <= 4 and (p.isdigit() or bool(re.match(r'^[A-Za-z]\d?$', p)))
        if not is_cas:
            return '.'.join(parts[i:])
    # 兜底：返回最后两段
    return '.'.join(parts[-2:]) if len(parts) >= 2 else code_str

def top_level_value(row, idx_mkt=-1, idx_cost=-1):
    """
    取一行的主值：
    优先市值 -> 成本 -> 行内最大数字
    """
    vals = [parse_num(x) for x in row]
    if idx_mkt != -1:
        v = parse_num(row[idx_mkt])
        if v != 0:
            return v
    if idx_cost != -1:
        v = parse_num(row[idx_cost])
        if v != 0:
            return v
    return max(vals) if vals else 0.0

def classify_asset(code, name):
    """
    基于 CAS 科目代码层级分类。
    优先级：非标特征字 > 明确科目代码 > 代码特征兜底。
    注意：code 参数会先经 clean_code_str 规范化（补交易所后缀点号），
    确保正则 \b 断词正确（如 301308.SZ 而非 301308SZ）。
    """
    c = clean_code_str(code)       # 用 clean_code_str 替代 norm_text，保证 .SH/.SZ 有点号
    n = norm_text(name)
    c_upper = c.upper()

    # 1) 非标优先
    if '信托' in n:
        return '信托计划'
    if '资管计划' in n or '资产管理计划' in n or '资产管理' in n:
        return '资管计划'
    if '基金' in n and '公司' not in n:
        return '公募/私募基金'

    # 2) 明确科目代码
    # 1102 及所有子类 → 股票（上交所 1102.01 / 深交所 1102.33 / 创业板 1102.34 等）
    if c.startswith('1102'):
        # 排除少数名称含"债"但代码误入 1102 的情形
        if '债' in n and '股票' not in n and '股' not in n:
            return '债券'
        return '股票'
    if c.startswith('1103'):
        return '债券'
    if c.startswith('1104') or '资产支持证券' in n or 'ABS' in c_upper:
        return '债券'
    if c.startswith('1105'):
        return 'ETF/基金投资'
    if c.startswith('1108'):
        return '其他交易性金融资产'
    if c.startswith('1201') or c.startswith('1202') or '买入返售' in n or '逆回购' in n or '质押式' in n:
        return '买入返售(逆回购)'

    # 3) 兜底判定（非标准代码）
    if re.search(r'\b(00|30|60|68|83|87|43|92)\d{4}\b', c) or '股' in n:
        return '股票'
    if '债' in n:
        return '债券'

    return '未分类'

def classify_position_bucket(code, name, asset_type=""):
    """
    新增仓位占比统计口径：
    现金类资产、信用债、转债、权益、指数类资产、其他。
    """
    c = clean_code_str(code)
    n = norm_text(name)
    t = norm_text(asset_type)
    n_upper = n.upper()

    if (
        c.startswith(('1002', '1021', '1031', '1201', '1202'))
        or t == '买入返售(逆回购)'
        or n in ['银行存款', '结算备付金', '存出保证金', '买入返售金融资产']
        or '逆回购' in n
        or '质押式' in n
    ):
        return '现金类资产'

    if c.startswith('1103.04') or '转债' in n or '可转债' in n:
        return '转债'

    if c.startswith(('1103', '1104')) or t == '债券' or '资产支持证券' in n:
        return '信用债'

    if c.startswith('1105') or t == 'ETF/基金投资' or 'ETF' in n_upper or '指数' in n or '中证' in n or '上证' in n or '创业板' in n:
        return '指数类资产'

    if c.startswith('1102') or t == '股票' or '股票' in n:
        return '权益'

    return '其他'

def safe_read_table(file):
    """读取 xls/xlsx/csv"""
    file_name = file.name.lower()
    if file_name.endswith('.csv'):
        encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb18030']
        last_err = None
        for enc in encodings:
            try:
                file.seek(0)
                return pd.read_csv(file, encoding=enc, header=None, dtype=str)
            except Exception as e:
                last_err = e
        raise last_err
    else:
        # xls 需要 xlrd；xlsx 用 openpyxl
        try:
            file.seek(0)
            return pd.read_excel(file, header=None, dtype=str, engine='openpyxl')
        except Exception:
            file.seek(0)
            return pd.read_excel(file, header=None, dtype=str, engine='xlrd')

# ================= 核心解析引擎 =================
def process_valuation_files(uploaded_files):
    summary_list = []
    detail_list = []

    for file in uploaded_files:
        file_name = os.path.basename(file.name)
        base_name = os.path.splitext(file_name)[0]
        product_name = re.sub(r'(_资产估值表.*|_四级.*)$', '', base_name)
        product_name = re.sub(r'^[A-Z0-9]+_', '', product_name)  # 去掉基金代码前缀如 SNK912_
        valuation_date = extract_valuation_date(file_name)

        # 1. 原始读取
        try:
            df = safe_read_table(file)
        except Exception as e:
            st.error(f"读取 {file_name} 失败: {str(e)}")
            continue

        # 2. 定位表头坐标
        header_idx = -1
        for i in range(min(30, len(df))):
            row_vals = [norm_text(x) for x in df.iloc[i].values]
            if '科目代码' in row_vals and '科目名称' in row_vals:
                header_idx = i
                break

        if header_idx == -1:
            st.warning(f"跳过 {file_name}：未能在前30行找到标准表头(科目代码/科目名称)")
            continue

        row0 = [norm_text(x) for x in df.iloc[header_idx].values]
        # 取第二行，但先判断是否为真正的双行表头（含 原币/本币 等二级标注）
        row1_raw = [norm_text(x) for x in df.iloc[header_idx + 1].values] if header_idx + 1 < len(df) else row0
        row1_is_header = any(k in row1_raw for k in ['本币', '原币', '金额', '数量'])
        row1 = row1_raw if row1_is_header else row0  # 非双行表头则退化为单行匹配

        # 动态寻找列索引
        def find_idx(keys0, keys1=None):
            if keys1:
                for i in range(len(row0)):
                    if any(k in row0[i] for k in keys0) and any(k in row1[i] for k in keys1):
                        return i
            for i in range(len(row0)):
                if any(k in row0[i] for k in keys0):
                    return i
            return -1

        idx_code = find_idx(['科目代码'])
        idx_name = find_idx(['科目名称'])
        idx_qty = find_idx(['数量'])
        idx_unit_cost = find_idx(['单位成本'])
        idx_price = find_idx(['行情', '市价'])
        idx_cost = find_idx(['成本'], ['本币'])
        idx_mkt = find_idx(['市值'], ['本币'])

        if idx_cost == -1:
            idx_cost = find_idx(['成本'])
        if idx_mkt == -1:
            idx_mkt = find_idx(['市值'])
        if idx_qty == -1:
            idx_qty = 4  # 常见数量列

        # 初始化统计桶
        total_assets = 0.0
        net_assets = 0.0
        bank_deposit = 0.0
        clearing_prov = 0.0
        margin_deposit = 0.0
        reverse_repo = 0.0
        reverse_repo_seen_top = False
        stocks = 0.0
        bonds = 0.0
        others = 0.0
        fund_etf = 0.0

        # 3. 逐行解析
        for i in range(header_idx + 1, len(df)):
            row = df.iloc[i].values

            # 防止极短行
            if len(row) <= max(idx_code if idx_code != -1 else 0, idx_name if idx_name != -1 else 0):
                continue

            c_raw = str(row[idx_code] if idx_code != -1 else "").strip()
            n_raw = str(row[idx_name] if idx_name != -1 else "").strip()

            c_clean = norm_text(c_raw)
            n_clean = norm_text(n_raw)

            # 结束标志：声明部分，后面不要再解析
            if n_clean.startswith('声明'):
                break

            # 空行跳过
            if not c_clean and not n_clean:
                continue

            row_val = top_level_value(row, idx_mkt, idx_cost)

            # ================= A. 顶层总计项（精确匹配，不能用 contains） =================
            # 资产合计 / 资产净值：只认精确行，防止“其中”与“声明”污染
            if c_clean == '资产合计' or n_clean == '资产合计':
                total_assets = row_val
                continue

            if c_clean == '资产净值' or n_clean == '资产净值':
                net_assets = row_val
                continue

            # 银行存款：你要求新增这一列
            # 这类通常是 1002 顶层行
            if c_clean == '1002' or n_clean == '银行存款':
                bank_deposit = row_val
                continue

            # 结算备付金：只取顶层 1021，避免把 1021.81、1021.B1 等子项重复加总
            if c_clean == '1021' or n_clean == '结算备付金':
                clearing_prov = row_val
                continue

            # 存出保证金：并入“结算备付金及保证金”口径
            if c_clean == '1031' or n_clean == '存出保证金':
                margin_deposit = row_val
                continue

            # 买入返售金融资产：你这份表是 1202，不是 1201
            if c_clean in ['1201', '1202'] or n_clean == '买入返售金融资产':
                reverse_repo = row_val
                reverse_repo_seen_top = True
                continue

            # ================= B. 底层持仓穿透（有数量才进入明细） =================
            v_qty = parse_num(row[idx_qty]) if idx_qty != -1 else 0.0

            # 只有真正有数量的明细才进入穿透统计
            if v_qty > 0:
                asset_type = classify_asset(c_raw, n_raw)

                # 排除明显汇总行、应计利息子行（避免与本金行重复计算）
                if any(x in n_clean for x in ['汇总', '合计', '大类', '交易所', '深交所', '上交所', '银行间', '小计', '其中', '应计利息']):
                    continue

                # 排除应计利息子行（科目代码含 .03. 层级标记）
                if is_accrued_interest_row(c_raw):
                    continue

                # 明细表：只记录资产类持仓
                if c_clean.startswith(('1102', '1103', '1104', '1105', '1108', '1201', '1202')):
                    clean_code = "逆回购" if asset_type == '买入返售(逆回购)' else extract_ticker(c_raw)

                    detail_list.append({
                        "所属产品": product_name,
                        "估值日期": valuation_date,
                        "文件名": file_name,
                        "资产类别": asset_type,
                        "代码": clean_code,
                        "资产名称": n_raw,
                        "数量": v_qty,
                        "单位成本": parse_num(row[idx_unit_cost]) if idx_unit_cost != -1 else 0.0,
                        "总成本": parse_num(row[idx_cost]) if idx_cost != -1 else 0.0,
                        "今日行情": parse_num(row[idx_price]) if idx_price != -1 else 0.0,
                        "今日市值": parse_num(row[idx_mkt]) if idx_mkt != -1 else 0.0
                    })

                # 汇总桶
                v_val = parse_num(row[idx_mkt]) if idx_mkt != -1 and parse_num(row[idx_mkt]) != 0 else parse_num(row[idx_cost])

                if asset_type == '股票':
                    stocks += v_val
                elif asset_type == '债券':
                    bonds += v_val
                elif asset_type == '买入返售(逆回购)':
                    if not reverse_repo_seen_top:
                        reverse_repo += v_val
                elif asset_type in ['信托计划', '资管计划', '公募/私募基金', '其他交易性金融资产']:
                    others += v_val
                elif asset_type == 'ETF/基金投资':
                    fund_etf += v_val

        # 单个产品解析完毕，压入汇总表
        summary_list.append({
            "产品名称": product_name,
            "估值日期": valuation_date,
            "文件名": file_name,
            "银行存款": bank_deposit,
            "结算备付金及保证金": clearing_prov + margin_deposit,
            "股票投资金额": stocks,
            "债券投资金额": bonds,
            "基金投资金额": fund_etf,
            "买入返售金融资产": reverse_repo,
            "其他交易性金融资产": others,
            "总资产": total_assets,
            "净资产": net_assets
        })

    df_sum = pd.DataFrame(summary_list)
    df_det = pd.DataFrame(detail_list)
    return df_sum, df_det

def build_position_allocation(df_sum, df_det):
    """生成新增 Sheet：按确认口径统计仓位占比，分母为总资产。"""
    columns = [
        "产品名称",
        "现金类资产占比",
        "信用债占比",
        "转债占比",
        "权益占比",
        "指数类资产占比",
        "其他",
        "未统计资产占比",
    ]
    if df_sum.empty:
        return pd.DataFrame(columns=columns)

    include_valuation_date = "估值日期" in df_sum.columns
    include_file_name = "文件名" in df_sum.columns
    if include_valuation_date:
        columns.insert(1, "估值日期")
    if include_file_name:
        insert_at = 2 if include_valuation_date else 1
        columns.insert(insert_at, "文件名")

    bucket_to_col = {
        "现金类资产": "现金类资产占比",
        "信用债": "信用债占比",
        "转债": "转债占比",
        "权益": "权益占比",
        "指数类资产": "指数类资产占比",
        "其他": "其他",
    }

    result_rows = []
    detail_df = df_det.copy() if df_det is not None and not df_det.empty else pd.DataFrame()

    for _, sum_row in df_sum.iterrows():
        product_name = sum_row.get("产品名称", "")
        valuation_date = sum_row.get("估值日期", "")
        file_name = sum_row.get("文件名", "")
        total_assets = parse_num(sum_row.get("总资产", 0))
        amounts = {bucket: 0.0 for bucket in bucket_to_col}

        # 现金类资产来自顶层科目，避免逆回购明细行重复计算。
        amounts["现金类资产"] += parse_num(sum_row.get("银行存款", 0))
        amounts["现金类资产"] += parse_num(sum_row.get("结算备付金及保证金", 0))
        amounts["现金类资产"] += parse_num(sum_row.get("买入返售金融资产", 0))

        if not detail_df.empty and "所属产品" in detail_df.columns:
            product_details = detail_df[detail_df["所属产品"] == product_name]
            if include_valuation_date and "估值日期" in detail_df.columns:
                product_details = product_details[product_details["估值日期"] == valuation_date]
            if include_file_name and "文件名" in detail_df.columns:
                product_details = product_details[product_details["文件名"] == file_name]
            for _, det_row in product_details.iterrows():
                bucket = classify_position_bucket(
                    det_row.get("代码", ""),
                    det_row.get("资产名称", ""),
                    det_row.get("资产类别", ""),
                )
                if bucket == "现金类资产":
                    continue
                amounts[bucket] += parse_num(det_row.get("今日市值", 0))

        out_row = {"产品名称": product_name}
        if include_valuation_date:
            out_row["估值日期"] = valuation_date
        if include_file_name:
            out_row["文件名"] = file_name
        classified_ratio = 0.0
        for bucket, col in bucket_to_col.items():
            out_row[col] = amounts[bucket] / total_assets if total_assets else 0.0
            classified_ratio += out_row[col]
        out_row["未统计资产占比"] = 1.0 - classified_ratio if total_assets else 0.0
        result_rows.append(out_row)

    return pd.DataFrame(result_rows, columns=columns)

def position_allocation_notes():
    """仓位占比统计 Sheet 底部备注说明。"""
    return [
        ("统计口径", "各项占比均以产品总资产为分母。"),
        ("现金类资产", "银行存款、结算备付金、存出保证金、买入返售金融资产/逆回购等。"),
        ("信用债", "普通债券、信用债、资产支持证券等；不含可转债。"),
        ("转债", "科目代码 1103.04 或名称包含“转债/可转债”的债券。"),
        ("权益", "股票类资产，包括沪深北交所股票、港股等。"),
        ("指数类资产", "ETF、指数基金，以及名称包含 ETF/指数/中证/上证/创业板等的资产。"),
        ("其他", "信托计划、资管计划、私募基金、其他交易性金融资产、违约债权等不穿透资产。"),
        ("未统计资产", "总资产中未归入上述六类的部分，通常包括应收利息、应收申购款、衍生品估值、其他应收应付轧差等。"),
    ]

def build_validation_report(df_sum, position_alloc_df):
    """生成校验/异常提示 Sheet。"""
    columns = [
        "产品名称",
        "估值日期",
        "文件名",
        "总资产",
        "已分类资产占比合计",
        "未统计资产占比",
        "异常提示",
    ]
    if df_sum.empty:
        return pd.DataFrame(columns=columns)

    ratio_cols = ["现金类资产占比", "信用债占比", "转债占比", "权益占比", "指数类资产占比", "其他"]
    key_cols = [col for col in ["产品名称", "估值日期", "文件名"] if col in df_sum.columns and col in position_alloc_df.columns]

    if key_cols:
        merged = df_sum.merge(position_alloc_df, on=key_cols, how="left", suffixes=("", "_仓位"))
    else:
        merged = df_sum.copy()
        for col in ratio_cols + ["未统计资产占比"]:
            merged[col] = position_alloc_df[col] if col in position_alloc_df.columns else 0.0

    rows = []
    for _, row in merged.iterrows():
        total_assets = parse_num(row.get("总资产", 0))
        classified_ratio = round(sum(parse_num(row.get(col, 0)) for col in ratio_cols), 10)
        uncovered_ratio = round(parse_num(row.get("未统计资产占比", 0)), 10)

        if total_assets == 0:
            message = "总资产缺失或为0"
        elif classified_ratio > 1.000001 or uncovered_ratio < -0.000001:
            message = f"分类占比超过100%，超出 {abs(uncovered_ratio):.2%}"
        elif uncovered_ratio > 0.0001:
            message = f"未统计资产占比 {uncovered_ratio:.2%}"
        else:
            message = "正常"

        rows.append({
            "产品名称": row.get("产品名称", ""),
            "估值日期": row.get("估值日期", ""),
            "文件名": row.get("文件名", ""),
            "总资产": total_assets,
            "已分类资产占比合计": classified_ratio,
            "未统计资产占比": uncovered_ratio,
            "异常提示": message,
        })

    return pd.DataFrame(rows, columns=columns)

def build_parse_summary(uploaded_files, df_sum, df_det, validation_df):
    """生成页面解析摘要。"""
    file_count = len(uploaded_files)
    product_count = df_sum["产品名称"].nunique() if "产品名称" in df_sum.columns else 0
    date_count = df_sum["估值日期"].nunique() if "估值日期" in df_sum.columns else 0
    detail_count = len(df_det)
    warning_count = 0
    if validation_df is not None and not validation_df.empty and "异常提示" in validation_df.columns:
        warning_count = int((validation_df["异常提示"] != "正常").sum())
    return (
        f"解析摘要：成功解析 {file_count} 个文件，覆盖 {product_count} 个产品、{date_count} 个估值日期，"
        f"底层明细 {detail_count} 条，校验提示 {warning_count} 条。"
    )

# ================= 跨产品合并分析引擎 =================
def build_cross_product_analysis(df_sum, df_det):
    """从产品汇总和明细表生成跨产品合并分析"""
    if df_sum.empty or df_det.empty:
        return None, None

    # ---- Sheet 3a: 跨产品重仓证券 ----
    # 仅保留有代码的证券（排除"逆回购"通用标记）
    det_with_code = df_det[df_det['代码'].notna() & (df_det['代码'] != '') & (df_det['代码'] != '逆回购')].copy()
    if det_with_code.empty:
        return None, None

    group_cols = []
    if '估值日期' in det_with_code.columns:
        group_cols.append('估值日期')
    group_cols.extend(['代码', '资产名称', '资产类别'])

    agg_kwargs = {
        '涉及产品数': ('所属产品', 'nunique'),
        '涉及产品': ('所属产品', lambda x: '、'.join(sorted(set(x)))),
        '合计数量': ('数量', 'sum'),
        '合计成本': ('总成本', 'sum'),
        '合计市值': ('今日市值', 'sum'),
        '各产品持仓明细': ('所属产品', lambda x: ' | '.join(
            f"{p}({v:,.0f}张, 市值{m:,.2f})" for p, v, m in sorted(
                set(zip(x, det_with_code.loc[x.index, '数量'], det_with_code.loc[x.index, '今日市值']))
            )
        )),
    }
    if '文件名' in det_with_code.columns:
        agg_kwargs['涉及文件'] = ('文件名', lambda x: '、'.join(sorted(set(x))))

    # 按估值日期+代码聚合：避免同一产品不同日期混在一起
    cross = det_with_code.groupby(group_cols).agg(**agg_kwargs).reset_index()

    # 按合计市值降序，跨产品持有的排前面
    cross['排序键'] = cross['涉及产品数'].apply(lambda x: 0 if x >= 2 else 1)
    sort_cols = ['排序键']
    ascending = [True]
    if '估值日期' in cross.columns:
        sort_cols.append('估值日期')
        ascending.append(False)
    sort_cols.append('合计市值')
    ascending.append(False)
    cross = cross.sort_values(sort_cols, ascending=ascending).drop(columns=['排序键'])
    cross = cross.reset_index(drop=True)

    # ---- Sheet 3b: 产品资产配置对比（仅占比，金额已在 Sheet 1） ----
    if not df_sum.empty:
        alloc = df_sum.copy()
        total_cols = ['银行存款', '结算备付金及保证金', '股票投资金额', '债券投资金额', '基金投资金额', '买入返售金融资产', '其他交易性金融资产']
        pct_cols = []
        for col in total_cols:
            if col in alloc.columns:
                pct_name = col.replace('投资金额', '').replace('金融资产', '')
                alloc[pct_name + '占比'] = alloc[col] / alloc['总资产'].replace(0, np.nan)
                pct_cols.append(pct_name + '占比')
        id_cols = [col for col in ['产品名称', '估值日期', '文件名'] if col in alloc.columns]
        alloc = alloc[id_cols + pct_cols]  # 仅保留标识列和占比列
    else:
        alloc = None

    return cross, alloc


# ================= 交互渲染 =================
uploaded_files = st.file_uploader(
    "📂 请上传估值表文件 (多选 Excel/CSV)",
    type=["xlsx", "xls", "csv"],
    accept_multiple_files=True
)

if uploaded_files:
    with st.spinner("引擎正在进行二维坐标矩阵重构与底层资产穿透..."):
        df_sum, df_det = process_valuation_files(uploaded_files)
        position_alloc_df = build_position_allocation(df_sum, df_det)
        validation_df = build_validation_report(df_sum, position_alloc_df)

    st.success(f"成功解析 {len(uploaded_files)} 个产品！")
    st.info(build_parse_summary(uploaded_files, df_sum, df_det, validation_df))

    # ---- 跨产品合并分析 ----
    cross_df, alloc_df = None, None
    if len(uploaded_files) >= 2:
        cross_df, alloc_df = build_cross_product_analysis(df_sum, df_det)

    # ================= Excel 导出 (先生成，下载按钮放最上面) =================
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_sum.to_excel(writer, sheet_name='产品概况', index=False)
        position_alloc_df.to_excel(writer, sheet_name='仓位占比统计', index=False)
        validation_df.to_excel(writer, sheet_name='校验提示', index=False)
        df_det.to_excel(writer, sheet_name='资产明细', index=False)
        if cross_df is not None and not cross_df.empty:
            cross_df.to_excel(writer, sheet_name='跨产品合并', index=False)
        if alloc_df is not None:
            alloc_df.to_excel(writer, sheet_name='产品配置对比', index=False)

        # ---- 格式美化 ----
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_font = Font(name='微软雅黑', size=10)
        cell_align = Alignment(horizontal='right', vertical='center')
        cell_align_left = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0'),
        )
        cross_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

        for ws in writer.sheets.values():
            ws.freeze_panes = 'A2'
            if ws.max_row > 1:
                ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.font = cell_font
                    cell.border = thin_border
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = cell_align
                        if abs(cell.value) >= 1000:
                            cell.number_format = '#,##0.00'
                        elif 0 < abs(cell.value) < 1:
                            cell.number_format = '0.00%' if cell.value <= 1 else '0.00'
                    elif cell.value is not None:
                        cell.alignment = cell_align_left

            if ws.title == '跨产品合并':
                cross_col_idx = None
                for col_idx, cell in enumerate(ws[1], 1):
                    if cell.value == '涉及产品数':
                        cross_col_idx = col_idx
                        break
                if cross_col_idx:
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        cell = row[cross_col_idx - 1]
                        try:
                            if cell.value is not None and float(cell.value) >= 2:
                                for c in row:
                                    c.fill = cross_fill
                        except (ValueError, TypeError):
                            pass

            if ws.title == '仓位占比统计':
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00%'

                note_start_row = ws.max_row + 2
                ws.cell(note_start_row, 1, "备注")
                ws.cell(note_start_row, 1).font = Font(name='微软雅黑', size=11, bold=True)
                ws.cell(note_start_row, 1).alignment = cell_align_left

                for offset, (bucket, note) in enumerate(position_allocation_notes(), start=1):
                    type_cell = ws.cell(note_start_row + offset, 1, bucket)
                    note_cell = ws.cell(note_start_row + offset, 2, note)
                    type_cell.font = Font(name='微软雅黑', size=10, bold=True)
                    type_cell.fill = PatternFill(start_color='D9EAF7', end_color='D9EAF7', fill_type='solid')
                    type_cell.alignment = cell_align_left
                    note_cell.font = cell_font
                    note_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    for c in (type_cell, note_cell):
                        c.border = thin_border

            if ws.title == '校验提示':
                for header_cell in ws[1]:
                    if '占比' in str(header_cell.value):
                        for cell in ws.iter_cols(
                            min_col=header_cell.column,
                            max_col=header_cell.column,
                            min_row=2,
                            max_row=ws.max_row,
                        ):
                            for c in cell:
                                if isinstance(c.value, (int, float)):
                                    c.number_format = '0.00%'

            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_width = 8
                for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            val = str(cell.value)
                            width = sum(2 if ord(c) > 127 else 1 for c in val)
                            max_width = max(max_width, min(width + 4, 55))
                ws.column_dimensions[col_letter].width = max_width

    st.download_button(
        label="📥 点击下载精算级汇总 Excel 报表",
        data=output.getvalue(),
        file_name=build_output_filename(uploaded_files),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # ================= 下方：数据预览 =================
    st.markdown("---")
    st.subheader("📋 Sheet 1: 产品资产概况")
    df_sum_show = df_sum.copy()
    for col in df_sum_show.columns:
        if col not in ["产品名称", "估值日期", "文件名"]:
            df_sum_show[col] = df_sum_show[col].apply(lambda x: f"{x:,.2f}")
    st.dataframe(df_sum_show, use_container_width=True)

    st.subheader("📊 Sheet 2: 仓位占比统计")
    position_alloc_show = position_alloc_df.copy()
    for col in position_alloc_show.columns:
        if col not in ["产品名称", "估值日期", "文件名"]:
            position_alloc_show[col] = position_alloc_show[col].apply(lambda x: f"{x:.2%}" if pd.notna(x) else "")
    st.dataframe(position_alloc_show, use_container_width=True)

    st.subheader("⚠️ Sheet 3: 校验提示")
    validation_show = validation_df.copy()
    for col in validation_show.columns:
        if '占比' in str(col):
            validation_show[col] = validation_show[col].apply(lambda x: f"{x:.2%}" if pd.notna(x) else "")
        elif col == "总资产":
            validation_show[col] = validation_show[col].apply(lambda x: f"{x:,.2f}")
    st.dataframe(validation_show, use_container_width=True)

    st.subheader("🔍 Sheet 4: 底层资产明细")
    df_det_show = df_det.copy()
    if not df_det_show.empty:
        for col in ['数量', '单位成本', '总成本', '今日行情', '今日市值']:
            if col in df_det_show.columns:
                df_det_show[col] = df_det_show[col].apply(lambda x: f"{x:,.2f}")
    st.dataframe(df_det_show, use_container_width=True)

    # ---- Sheet 3: 跨产品合并分析 ----
    if cross_df is not None and not cross_df.empty:
        st.subheader("🔗 Sheet 5: 跨产品合并分析")

        st.markdown("**跨产品重仓证券**（同一只券被多个产品持有，合并计算敞口）")
        cross_show = cross_df.copy()
        for col in ['合计数量', '合计成本', '合计市值']:
            if col in cross_show.columns:
                cross_show[col] = cross_show[col].apply(lambda x: f"{x:,.2f}")
        st.dataframe(cross_show, use_container_width=True)

        if alloc_df is not None:
            st.markdown("**产品资产配置对比**")
            alloc_show = alloc_df.copy()
            for col in alloc_show.columns:
                if '占比' in str(col):
                    alloc_show[col] = alloc_show[col].apply(lambda x: f"{x:.2%}" if pd.notna(x) else "")
                elif col not in ["产品名称", "估值日期", "文件名"]:
                    alloc_show[col] = alloc_show[col].apply(lambda x: f"{x:,.2f}")
            st.dataframe(alloc_show, use_container_width=True)
    elif len(uploaded_files) >= 2:
        st.info("当前产品之间无共同持仓，或明细数据不足以进行跨产品分析。")

else:
    st.info("提示：支持批量拖入估值表。算法已升级为精确总计行识别 + 底层明细穿透，避免子项重复累计。")
